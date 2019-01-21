import openpyxl
import datetime
from openpyxl.styles.borders import Border, Side
from PIL import Image

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"
sheet1.cell(row=1, column=1).value = 'Title'
sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
sheet2['A1'] = datetime.datetime.now()
sheet2['A2'] = datetime.date.today()
# 저장하기


sheet1['C3'].value = " "
sheet1['B2'].value = 123123123
tmpCell = sheet1['C3']
# tmpCell = sheet1.cell(row=1, column=3)
tmpCell.font = openpyxl.styles.Font(size=40, color='FF0000')
tmpCell.number_format


# insert image
imgFile = '../crawl/images/1.png'
img = openpyxl.drawing.image.Image(imgFile)
sheet2.add_image(img, 'C3')
# resize image

img2 = Image.open(imgFile)
new_img = img2.resize((73, 77))
new_img.save('new.png')
img3 = openpyxl.drawing.image.Image('new.png')
sheet1.add_image(img3, 'C3')


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='none'), 
                     bottom=Side(style='thin'))

cell = sheet1['B3']
cell.border = thin_border






book.save("./data/output.xlsx")