from bs4 import BeautifulSoup
import requests
import csv
import codecs
import time
import openpyxl
from PIL import Image
from openpyxl.chart import (
    Reference,
    BarChart)
from openpyxl.chart import (
    Reference, Series,
    ScatterChart
)
# ===============================================================================
fp = codecs.open("melon_add_likecnt.csv", "r", encoding="utf-8")
reader = csv.reader(fp, delimiter=',', quotechar='"')

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"

column = "ABCDE"
for j, i in enumerate(reader):
    for num, data in enumerate(i):
        row_num = j+1
        if row_num == 1 :
            sheet1[str(column[num] + str(row_num))] = data
            
        else : 
            if num == 0 or num == 3 or num == 4 :
                if data == '총계' :
                    data = data
                    sheet1[str(column[num] + str(row_num))] = data
                else :
                    data = int(data)
                    sheet1[str(column[num] + str(row_num))] = data
            else :
                sheet1[str(column[num] + str(row_num))] = data
#====================================================

url = "https://www.melon.com/chart/index.htm"

headers = {
    'Host': 'www.melon.com',
    'If-None-Match' : "0:b170",
    'Referer': 'https://www.melon.com/index.htm',
    'Upgrade-Insecure-Requests':'1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36'
}

html = requests.get(url, headers = headers).text
singer = []
rank = []
song_name = []
sheet_2 = []

soup = BeautifulSoup(html, 'html.parser')
sel_song = "#frm table tbody tr "

get_song = soup.select(sel_song)
for i in get_song:
    rank.append(i.select_one('div span.rank').text)
    song_name.append((i.select_one('div.rank01 span a').text))
    singer.append(i.select_one('div.rank02 span').text)

sel_img = "#frm table tbody tr div.wrap a img"
get_img = soup.select(sel_img)
for num, j in enumerate(get_img):
    src = j.get('src')
    img_src = requests.get(src).content
    saveFile = str(rank[num]) + ".png"
    with open("./images/" +saveFile, mode="wb") as file:
        file.write(img_src)

sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
col_num = 0
for i in range(100):
    j = i+1
    tmpCell_1 = sheet2[str(column[3] + str(j))]
    tmpCell_1.font = openpyxl.styles.Font(size=50, color='FF0000')
    sheet2[str(column[0] + str(j))] = rank[i]
    sheet2[str(column[1] + str(j))] = song_name[i]
    sheet2[str(column[2] + str(j))] = singer[i]
    sheet2[str(column[3] + str(j))] = "  "
    imgFile = './images/{}.png'.format(j)
    img2 = Image.open(imgFile)
    new_img = img2.resize((70, 83))
    new_img.save('./images/new{}.png'.format(j))
    img2 = openpyxl.drawing.image.Image('./images/new{}.png'.format(j))
    sheet2.add_image(img2, '{}'.format(column[4] + str(j)))
    
#====================================================

sheet3 = book.create_sheet()
sheet3.title = "세번째 시트"
datax = Reference(sheet1, min_col=4, 
		min_row=2, max_col=4, max_row=11)
categs = Reference(sheet1, min_col=2,
				 min_row=2, max_row=11)
chart = BarChart()
chart.add_data(data=datax)
chart.set_categories(categs)
chart.legend = None  # 범례
chart.varyColors = True
chart.title = "좋아요"
sheet3.add_chart(chart, "A1")

chart = ScatterChart()
chart.style = 13
chart.x_axis.title = 'Ranking'
chart.y_axis.title = '좋아요 차이'

xvalues = Reference(sheet1, min_col=1,
			 min_row=2, max_row=11)
values = Reference(sheet1, 
            min_col=5, 
            min_row=1, 
            max_row=11)
series = Series(values, xvalues, 
            title_from_data=True)
chart.series.append(series)
sheet3.add_chart(chart, "A20")


book.save("./meltop.xlsx")
# # insert image







